"""Generate JSON resources for the Log Adviser RuneLite plugin from the
Collection Log Adviser spreadsheet.

Inputs:
  Collection Log Adviser.xlsx (in the project root)
  tools/category_overrides.json

Outputs (under src/main/resources/com/logadviser/data/):
  activities.json      - one entry per row of "List of activities"
  activity_map.json    - one entry per row of "Activity map" (preserves row
                         order so "requires previous" semantics survive the
                         JSON trip)
  slots.json           - one entry per row of "List of log slots"
"""

from __future__ import annotations

import json
import os
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "Collection Log Adviser.xlsx"
OVERRIDES = ROOT / "tools" / "category_overrides.json"
OUT_DIR = ROOT / "src" / "main" / "resources" / "com" / "logadviser" / "data"

MINIGAME_KEYWORDS = [
    "pest control", "castle wars", "trouble brewing", "soul wars",
    "last man standing", "mahogany homes", "hallowed sepulchre",
    "temple trekking", "barbarian assault", "gnome restaurant",
    "nightmare zone", "tithe farm", "volcanic mine", "giants' foundry",
    "guardians of the rift", "tempoross", "wintertodt", "fishing trawler",
    "pyramid plunder", "rogues' den", "gauntlet", "inferno", "fight caves",
    "agility arena", "brimhaven vouchers", "shades of mort'ton",
    "mage training arena", "magic training arena", "blast furnace",
    "motherlode", "mastering mixology", "mixology", "fortis colosseum",
    "barracuda trials", "vale totems", "vale offerings", "sailing islands",
    "port tasks", "ocean encounters", "hunters' rumours",
    "moons of peril", "doom of mokhaiotl", "salvage", "stronghold of security",
    "creature creation",
]

COMBAT_VERBS = ("killing ", "defeating ", "fighting ", "delving to depth ")


def categorize(name: str) -> str:
    n = name.lower()
    if "(boat combat)" in n:
        return "combat"
    if n.startswith("theater of blood") or n.startswith("ancient chest (chambers"):
        return "combat"
    if n.startswith("tombs of amascut") or n.startswith("completing tombs of amascut"):
        return "combat"
    if "(champion's challenge)" in n:
        return "combat"
    if "(royal titans)" in n:
        return "combat"
    for kw in MINIGAME_KEYWORDS:
        if kw in n:
            return "minigame"
    if any(n.startswith(v) for v in COMBAT_VERBS):
        return "combat"
    if n.startswith("looting "):
        return "combat"
    return "miscellaneous"


def coerce_int(v):
    if v is None or v == "":
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def coerce_float(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def coerce_bool(v):
    return bool(v) if v is not None else False


def load_overrides() -> dict[int, str]:
    if not OVERRIDES.exists():
        return {}
    raw = json.loads(OVERRIDES.read_text(encoding="utf-8"))
    return {int(k): v for k, v in raw.items() if not k.startswith("_")}


def write_json(path: Path, payload) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    text = json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=False)
    path.write_text(text + "\n", encoding="utf-8")
    print(f"  wrote {path.relative_to(ROOT)} ({len(payload)} entries)")


def build_completion_rates(wb) -> dict[int, dict]:
    ws = wb["Completion rates"]
    rates: dict[int, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        idx = coerce_int(row[0])
        if idx is None:
            continue
        rates[idx] = {
            "completionsPerHrMain": coerce_float(row[2]) or 0.0,
            "completionsPerHrIron": coerce_float(row[3]) or 0.0,
            "extraTimeFirst": coerce_float(row[4]) or 0.0,
        }
    return rates


def build_activities(wb, rates, overrides) -> list[dict]:
    ws = wb["List of activities"]
    out = []
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        idx = coerce_int(row[0])
        name = row[1]
        if idx is None or not name or idx in seen:
            continue
        seen.add(idx)
        rate = rates.get(idx, {})
        category = overrides.get(idx) or categorize(name)
        out.append({
            "index": idx,
            "name": name,
            "completionsPerHrMain": rate.get("completionsPerHrMain", 0.0),
            "completionsPerHrIron": rate.get("completionsPerHrIron", 0.0),
            "extraTimeFirst": rate.get("extraTimeFirst", 0.0),
            "category": category,
        })
    out.sort(key=lambda a: a["index"])
    return out


def build_activity_map(wb) -> list[dict]:
    """Preserves the sheet's row order. Order matters because the
    `requiresPrevious` flag points at the *previous row of the same activity*."""
    ws = wb["Activity map"]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        idx = coerce_int(row[0])
        item_id = coerce_int(row[2])
        if idx is None or item_id is None:
            continue
        out.append({
            "activityIndex": idx,
            "itemId": item_id,
            "itemName": row[3],
            "slotDifficulty": coerce_int(row[4]) or 0,
            "requiresPrevious": coerce_bool(row[6]),
            "exact": coerce_bool(row[8]),
            "independent": coerce_bool(row[9]),
            "dropRateAttempts": coerce_float(row[10]) or 0.0,
        })
    return out


def build_slots(wb) -> list[dict]:
    ws = wb["List of log slots"]
    out = []
    seen = set()
    for row in ws.iter_rows(min_row=3, values_only=True):  # row 2 is a stats row
        item_id = coerce_int(row[0])
        name = row[1]
        if item_id is None or not name or item_id in seen:
            continue
        seen.add(item_id)
        out.append({
            "itemId": item_id,
            "slotName": name,
            "activityCount": coerce_int(row[5]) or 0,
        })
    out.sort(key=lambda s: s["itemId"])
    return out


def main() -> int:
    if not XLSX.exists():
        print(f"ERROR: {XLSX} not found", file=sys.stderr)
        return 1
    print(f"Reading {XLSX.name}")
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    overrides = load_overrides()
    rates = build_completion_rates(wb)
    activities = build_activities(wb, rates, overrides)
    activity_map = build_activity_map(wb)
    slots = build_slots(wb)

    print(f"  activities:     {len(activities)}")
    print(f"  activity_map:   {len(activity_map)}")
    print(f"  slots:          {len(slots)}")
    print(f"  overrides used: {len(overrides)}")

    # Sanity check: every activityIndex in activity_map should exist in activities
    activity_ids = {a["index"] for a in activities}
    map_ids = {row["activityIndex"] for row in activity_map}
    orphans = map_ids - activity_ids
    if orphans:
        print(f"WARNING: activity_map references unknown indices: {sorted(orphans)}")

    # Sanity check: every itemId in activity_map should exist in slots
    slot_ids = {s["itemId"] for s in slots}
    item_ids = {row["itemId"] for row in activity_map}
    missing = item_ids - slot_ids
    if missing:
        print(f"WARNING: activity_map references unknown item IDs: {len(missing)} missing")

    # Category counts
    counts: dict[str, int] = {}
    for a in activities:
        counts[a["category"]] = counts.get(a["category"], 0) + 1
    print(f"  category counts: {counts}")

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    write_json(OUT_DIR / "activities.json", activities)
    write_json(OUT_DIR / "activity_map.json", activity_map)
    write_json(OUT_DIR / "slots.json", slots)
    print("Done.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
